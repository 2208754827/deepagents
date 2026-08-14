#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试用例 XLSX 生成器
生成格式美观的 Excel 文件，直接保存到系统桌面

依赖：pip install openpyxl
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误：请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── 常量定义 ─────────────────────────────────────────────────────────────────

HEADERS = [
    "模块",
    "用例标题",
    "优先级",
    "需求ID",
    "设计方法",
    "前置条件",
    "测试步骤",
    "预期结果",
    "备注",
    "标签",
]

# 列宽配置
COL_WIDTHS = {
    'A': 20,  # 模块
    'B': 35,  # 用例标题
    'C': 10,  # 优先级
    'D': 15,  # 需求ID
    'E': 12,  # 设计方法
    'F': 25,  # 前置条件
    'G': 50,  # 测试步骤
    'H': 45,  # 预期结果
    'I': 20,  # 备注
    'J': 15,  # 标签
}

# 优先级颜色
PRIORITY_COLORS = {
    'P0': 'FF0000',  # 红色
    'P1': 'FF6600',  # 橙色
    'P2': 'FFCC00',  # 黄色
    'P3': '99CC00',  # 绿色
}


# ── 数据格式化 ────────────────────────────────────────────────────────────────

def _normalize_text(value) -> str:
    """把 JSON 里的值统一转成纯文本。"""
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        parts = [_normalize_text(item) for item in value]
        return ", ".join(part for part in parts if part)
    return str(value).strip()


def _pick_value(step: dict, keys: tuple) -> str:
    """从步骤字典中按优先级取第一个非空值。"""
    if not isinstance(step, dict):
        return ""
    for key in keys:
        val = _normalize_text(step.get(key))
        if val:
            return val
    return ""


def _strip_leading_number(text: str) -> str:
    """去掉文本开头的序号前缀"""
    return re.sub(r"^\d+\.\s*", "", text.strip(), count=1)


def _format_actions(steps) -> str:
    """仅提取操作列"""
    if not isinstance(steps, list):
        return _normalize_text(steps)

    lines = []
    for idx, step in enumerate(steps, start=1):
        if isinstance(step, dict):
            action = _pick_value(step, ("操作", "动作", "步骤"))
            if action:
                clean = _strip_leading_number(action)
                lines.append(f"{idx}. {clean}")
            else:
                lines.append(f"{idx}.")
        else:
            text = _normalize_text(step)
            if text:
                lines.append(f"{idx}. {text}")
    return "; ".join(lines)


def _format_expected(steps) -> str:
    """仅提取预期结果列"""
    if not isinstance(steps, list):
        return ""

    lines = []
    for idx, step in enumerate(steps, start=1):
        if isinstance(step, dict):
            expected = _pick_value(step, ("预期", "预期结果", "期望"))
            if expected:
                clean = _strip_leading_number(expected)
                lines.append(f"{idx}. {clean}")
            else:
                lines.append(f"{idx}.")
        else:
            lines.append(f"{idx}.")
    return "; ".join(lines)


def _case_to_row(case: dict) -> dict:
    """把单条用例 JSON 映射为表格行"""
    module_path = case.get("模块") or []
    if isinstance(module_path, str):
        module_path = [module_path]

    steps = case.get("步骤") or []

    return {
        "模块": " / ".join(_normalize_text(item) for item in module_path if _normalize_text(item)),
        "用例标题": _normalize_text(case.get("用例标题") or case.get("标题")),
        "优先级": _normalize_text(case.get("优先级") or case.get("priority")),
        "需求ID": _normalize_text(case.get("需求ID") or case.get("req_id")),
        "设计方法": _normalize_text(case.get("设计方法") or case.get("method")),
        "前置条件": _normalize_text(case.get("前置条件") or case.get("precondition")),
        "测试步骤": _format_actions(steps),
        "预期结果": _format_expected(steps),
        "备注": _normalize_text(case.get("备注") or case.get("remark")),
        "标签": _normalize_text(case.get("标签") or case.get("tag")),
    }


# ── XLSX 生成 ────────────────────────────────────────────────────────────────

def _build_styles():
    """构建 XLSX 样式"""
    return {
        'header_font': Font(bold=True, color="FFFFFF", size=11),
        'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
        'header_align': Alignment(horizontal='center', vertical='center'),
        'cell_align': Alignment(vertical='top', wrap_text=True),
    }


def generate_xlsx(cases: list, output_path: str):
    """生成格式美观的 XLSX 文件"""
    output_path = Path(output_path)
    xlsx_path = output_path if output_path.suffix == '.xlsx' else output_path.with_suffix('.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    style = _build_styles()

    # 写入表头
    for col, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = style['header_font']
        cell.fill = style['header_fill']
        cell.alignment = style['header_align']
        cell.border = style['border']

    # 设置列宽
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 写入数据
    for row_idx, case in enumerate(cases, start=2):
        row_data = _case_to_row(case)
        for col_idx, header in enumerate(HEADERS, 1):
            value = row_data.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = style['cell_align']
            cell.border = style['border']

            # 优先级列着色
            if header == "优先级" and value in PRIORITY_COLORS:
                color = PRIORITY_COLORS[value]
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                if value in ['P0', 'P1']:
                    cell.font = Font(bold=True, color="FFFFFF")

    # 首行行高、冻结首行
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = 'A2'

    wb.save(xlsx_path)
    print(f"已生成: {xlsx_path.resolve()}")


# ── CLI 入口 ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="测试用例 XLSX 生成器")
    parser.add_argument("-o", "--output", required=True, help="输出 .xlsx 文件路径")
    parser.add_argument("-d", "--data", help="JSON 格式测试用例数据（字符串）")
    parser.add_argument("-f", "--file", help="JSON 文件路径")
    args = parser.parse_args()

    try:
        if args.file:
            cases = json.loads(Path(args.file).read_text(encoding="utf-8-sig"))
        elif args.data:
            cases = json.loads(args.data)
        else:
            cases = json.loads(sys.stdin.read())

        if not isinstance(cases, list):
            raise ValueError("输入数据必须是 JSON 数组")

        generate_xlsx(cases, args.output)

    except json.JSONDecodeError as e:
        print(f"JSON 解析错误: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"数据错误: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"生成失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
